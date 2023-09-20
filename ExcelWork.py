import openpyxl

def create_excel():
    workbook = openpyxl.Workbook()
    return workbook

def create_sheet(workbook, sheet_name):
    sheet = workbook.create_sheet(sheet_name)
    return sheet

def create_row(sheet, data):
    sheet.append(data)

def create_column(sheet, data):
    for value in data:
        sheet.append([value])

def navigate_to_cell(sheet, row, column):
    return sheet.cell(row=row, column=column)

def perform_formula(sheet, formula):
    sheet.append(["Formula Result"])
    cell = sheet.cell(row=sheet.max_row, column=1)
    cell.formula = formula

def show_excel_data(sheet):
    print("Sr\tValue")
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            print(f"{cell.row}\t{cell.value}")

def save_workbook(workbook, filename):
    workbook.save(filename)

def open_workbook(filename):
    return openpyxl.load_workbook(filename)

def main():
    while True:
        print("Options:")
        print("1. Create Sheets")
        print("2. Open File")
        print("3. Exit")
        
        choice = input("Enter your choice: ")

        if choice == "1":
            workbook = create_excel()
            sheet_name = input("Enter sheet name: ")
            sheet = create_sheet(workbook, sheet_name)
            
            while True:
                print("Sheet Options:")
                print("1. Create Row")
                print("2. Create Column")
                print("3. Navigate to Cell")
                print("4. Perform Formula")
                print("5. Show Excel Data")
                print("6. Save File")
                print("7. Back to Main Menu")
                
                sheet_choice = input("Enter your choice: ")
                
                if sheet_choice == "1":
                    row_data = input("Enter row data (comma-separated): ").split(',')
                    create_row(sheet, row_data)
                elif sheet_choice == "2":
                    column_data = input("Enter column data (comma-separated): ").split(',')
                    create_column(sheet, column_data)
                elif sheet_choice == "3":
                    row = int(input("Enter row number: "))
                    column = int(input("Enter column number: "))
                    cell = navigate_to_cell(sheet, row, column)
                    print(f"Value at ({row}, {column}): {cell.value}")
                elif sheet_choice == "4":
                    formula = input("Enter formula (e.g., SUM(A1:A3)): ")
                    perform_formula(sheet, formula)
                elif sheet_choice == "5":
                    show_excel_data(sheet)
                elif sheet_choice == "6":
                    filename = input("Enter file name to save: ")
                    save_workbook(workbook, filename)
                    print(f"File '{filename}' saved.")
                elif sheet_choice == "7":
                    break
                else:
                    print("Invalid choice. Please try again.")
                    
        elif choice == "2":
            filename = input("Enter file name to open: ")
            workbook = open_workbook(filename)
            print(f"File '{filename}' opened.")
        elif choice == "3":
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
